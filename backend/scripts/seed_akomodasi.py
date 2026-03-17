"""
Seed script: membaca dparis_akomodasi.xlsx → mengisi tabel accommodations di SQLite
Jalankan dari folder backend/: python scripts/seed_akomodasi.py
"""
import os
import sys

# Tambahkan root backend ke sys.path agar bisa import app.*
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl

from app.db.session import engine, SessionLocal
from app.models.base import Base
from app.models.accommodation import Accommodation  # noqa
import app.db.base  # noqa

XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "dparis_akomodasi.xlsx")

# Kolom Excel (0-based index sesuai header):
# 0:NO, 1:NAMA, 2:RATING, 3:ALAMAT, 4:TELEPON, 5:JENIS AKOMODASI,
# 6:KAMAR, 7:TEMPAT TIDUR, 8:AIR PANAS DINGIN, 9:TV KABEL, 10:FREE WIFI,
# 11:RESTORAN, 12:KOLAM_RENANG, 13:KEBUGARAN, 14:RUANG_MEETING, 15:LINK,
# 16:Jarak ke GUNUNG DEMPO, 17:Jarak ke PASAR DEMPO PERMAI,
# 18:Jarak ke BANDARA ATUNG BUNGSU, 19:Jarak ke RSUD BESEMAH,
# 20:Jarak ke SPBU AIR PERIKAN, 21:Jarak ke SPBU SIMPANG MANNA,
# 22:Jarak ke SPBU PENGANDONAN, 23:Jarak ke SPBU KARANG DALO


def ada(val) -> bool:
    """Konversi 'Ada'/'Tidak Ada' ke boolean."""
    return str(val).strip().lower() == "ada" if val else False


def to_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def clean_phone(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if s.isdigit() and len(s) >= 9:
        return "0" + s if not s.startswith("0") else s
    return s


def seed():
    xlsx_abs = os.path.abspath(XLSX_PATH)
    if not os.path.exists(xlsx_abs):
        print(f"[ERROR] File tidak ditemukan: {xlsx_abs}")
        sys.exit(1)

    # Pastikan tabel ada (dengan kolom baru)
    Base.metadata.create_all(bind=engine)

    wb = openpyxl.load_workbook(xlsx_abs)

    # Coba baca dari sheet aktif
    ws = wb.active

    # Baca header baris pertama untuk validasi urutan kolom
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"[INFO] Header Excel: {headers[:16]}")

    db = SessionLocal()
    try:
        # Hapus data lama
        deleted = db.query(Accommodation).delete()
        db.commit()
        print(f"[INFO] Hapus {deleted} data akomodasi lama")

        inserted = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue

            def g(idx, default=None):
                """Get value safely by index."""
                return row[idx] if len(row) > idx else default

            no          = int(g(0, 0))
            name        = str(g(1, "")).strip()
            rating      = to_float(g(2))
            address     = str(g(3, "")).strip() if g(3) else ""
            phone       = clean_phone(g(4))
            category    = str(g(5, "")).strip() if g(5) else "Akomodasi"
            rooms       = int(g(6, 0)) if g(6) else 0
            beds        = int(g(7, 0)) if g(7) else 0
            hot_water   = ada(g(8))
            tv_cable    = ada(g(9))
            free_wifi   = ada(g(10))
            restaurant  = ada(g(11))
            swimming    = ada(g(12))
            gym         = ada(g(13))
            meeting     = ada(g(14))
            link        = str(g(15, "")).strip() if g(15) else ""

            # Distance columns (km) — may be None if not yet calculated
            dist_gunung_dempo         = to_float(g(16))
            dist_pasar_dempo_permai   = to_float(g(17))
            dist_bandara_atung_bungsu = to_float(g(18))
            dist_rsud_besemah         = to_float(g(19))
            dist_spbu_air_perikan     = to_float(g(20))
            dist_spbu_simpang_manna   = to_float(g(21))
            dist_spbu_pengandonan     = to_float(g(22))
            dist_spbu_karang_dalo     = to_float(g(23))

            if not name:
                continue

            obj = Accommodation(
                id=no,
                name=name,
                rating=rating,
                address=address or None,
                phone=phone or None,
                category=category or "Akomodasi",
                total_rooms=rooms,
                total_beds=beds,
                hot_water=hot_water,
                tv_cable=tv_cable,
                free_wifi=free_wifi,
                restaurant=restaurant,
                swimming_pool=swimming,
                gym=gym,
                meeting_room=meeting,
                maps_link=link or None,
                dist_gunung_dempo=dist_gunung_dempo,
                dist_pasar_dempo_permai=dist_pasar_dempo_permai,
                dist_bandara_atung_bungsu=dist_bandara_atung_bungsu,
                dist_rsud_besemah=dist_rsud_besemah,
                dist_spbu_air_perikan=dist_spbu_air_perikan,
                dist_spbu_simpang_manna=dist_spbu_simpang_manna,
                dist_spbu_pengandonan=dist_spbu_pengandonan,
                dist_spbu_karang_dalo=dist_spbu_karang_dalo,
                is_active=True,
            )
            db.add(obj)
            inserted += 1

        db.commit()
        print(f"[OK] {inserted} data akomodasi berhasil disimpan ke DB (tabel: accommodations)")
    except Exception as e:
        db.rollback()
        print(f"[ERROR] Error: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    seed()
