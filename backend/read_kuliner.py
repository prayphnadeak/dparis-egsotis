import openpyxl

wb = openpyxl.load_workbook(r'c:\xampp\htdocs\dparis\dparis_kuliner.xlsx')
out = []
for sh in wb.sheetnames:
    out.append(f'=== SHEET: {sh} ===')
    ws = wb[sh]
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row):
            out.append(repr(list(row)))

with open(r'c:\xampp\htdocs\dparis\kuliner_dump.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(out))

print('Done, lines:', len(out))
