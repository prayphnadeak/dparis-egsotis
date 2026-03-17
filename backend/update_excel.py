import openpyxl
import re
import math
import shutil

def haversine(lat1, lon1, lat2, lon2):
    R = 6371000
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    
    a = math.sin(delta_phi / 2.0) ** 2 + \
        math.cos(phi1) * math.cos(phi2) * \
        math.sin(delta_lambda / 2.0) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return int(round(R * c))

def extract_coords(url):
    if not isinstance(url, str) or not url.strip():
        return None, None
    match_3d4d = re.search(r'!3d(-?\d+\.\d+)!4d(-?\d+\.\d+)', url)
    if match_3d4d:
        return float(match_3d4d.group(1)), float(match_3d4d.group(2))
    
    match_at = re.search(r'/@(-?\d+\.\d+),(-?\d+\.\d+)', url)
    if match_at:
        return float(match_at.group(1)), float(match_at.group(2))
    return None, None

file_path = 'dparis_akomodasi.xlsx'
backup_path = 'dparis_akomodasi_backup.xlsx'

# Backup original
shutil.copy2(file_path, backup_path)

wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Find LINK column
link_col_idx = None
headers = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val:
        headers[val] = col
        if str(val).strip().upper() == 'LINK':
            link_col_idx = col

if not link_col_idx:
    print("Error: LINK column not found.")
    exit(1)

# Check if JARAK KE GUNUNG DEMPO already exists
jarak_col_idx = None
if 'JARAK KE GUNUNG DEMPO' in headers:
    jarak_col_idx = headers['JARAK KE GUNUNG DEMPO']
else:
    jarak_col_idx = ws.max_column + 1
    ws.cell(row=1, column=jarak_col_idx, value='JARAK KE GUNUNG DEMPO')

gunung_dempo = (-4.0158333, 103.1283333)
success = 0
failed_but_not_empty = 0

for row in range(2, ws.max_row + 1):
    link_val = ws.cell(row=row, column=link_col_idx).value
    if link_val:
        lat, lon = extract_coords(str(link_val))
        if lat is not None and lon is not None:
            dist = haversine(lat, lon, gunung_dempo[0], gunung_dempo[1])
            ws.cell(row=row, column=jarak_col_idx, value=dist)
            success += 1
        else:
            print(f"Failed to parse coordinates from link at row {row}: {link_val}")
            failed_but_not_empty += 1
    else:
        # Leave distance empty if link is empty
        pass

if failed_but_not_empty == 0:
    wb.save(file_path)
    print(f"Successfully processed {success} coordinates and saved to {file_path}")
else:
    print(f"Failed to process {failed_but_not_empty} non-empty links. File not saved.")
